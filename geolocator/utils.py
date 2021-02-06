import re
import threading
from django.conf import settings
import openpyxl
from geopy.geocoders import MapQuest

def validate_address(address):
    if address!= None and type(address) == str:
        regex = re.compile("\d+(,?)[ ](?:[A-Za-z0-9.,-]+[ ]?)+")
        if len(regex.findall(address)) == 0:
            return False
        else:
            return True

def get_addresses(wb,ws,output):
    worksheet = wb[ws]
    for row in worksheet.iter_rows():
        for cell in row:
            if validate_address(cell.value):
                output.append(cell.value)

def parse_excel(excel_file):
        output = []
        wb = openpyxl.load_workbook(excel_file)
        worksheets = [i for i in wb.sheetnames]
        geolocator = MapQuest(api_key=settings.MAPQUEST_APIKEY)
        threads = []
        for worksheet in worksheets:
            t = threading.Thread(target=get_addresses,args=(wb,worksheet,output,))
            t.start()
            threads.append(t)
        for thread in threads:
            thread.join()    
        output = list(set(output))
        output = [[i,geolocator.geocode(i).latitude,geolocator.geocode(i).longitude] for i in output]    
        return output


